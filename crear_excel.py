from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Tipografía Boulevard"

# Ancho de columnas
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 35

# Estilos
header_fill = PatternFill(start_color="0671AE", end_color="0671AE", fill_type="solid")
header_font = Font(name="Nunito Sans", size=11, bold=True, color="FFFFFF")
title_fill = PatternFill(start_color="84CE25", end_color="84CE25", fill_type="solid")
title_font = Font(name="Nunito Sans", size=12, bold=True, color="FFFFFF")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# TÍTULO
ws['A1'] = "TIPOGRAFÍA DE BOULEVARD SANTA ROSA - EXPLICACIÓN SIMPLE"
ws['A1'].font = Font(name="Nunito Sans", size=14, bold=True, color="FFFFFF")
ws['A1'].fill = PatternFill(start_color="033D6B", end_color="033D6B", fill_type="solid")
ws['A1'].alignment = center_align
ws.merge_cells('A1:E1')
ws.row_dimensions[1].height = 25

# INFO GENERAL
ws['A3'] = "Fuente Principal:"
ws['B3'] = "Nunito Sans"
ws['A4'] = "Dónde viene:"
ws['B4'] = "Google Fonts (gratuita)"
ws['A5'] = "Cantidad de letras diferentes:"
ws['B5'] = "SOLO UNA (Nunito Sans)"

for row in [3, 4, 5]:
    ws[f'A{row}'].font = Font(bold=True, size=10)
    ws[f'B{row}'].font = Font(size=10)
    ws.row_dimensions[row].height = 18

# ENCABEZADOS TABLA
headers = ["QUE VES EN LA PAGINA", "COMO SE LLAMA", "QUE TAN GRANDE", "QUE TAN GRUESA", "DONDE LO VES"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=7, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

ws.row_dimensions[7].height = 25

# DATOS
data = [
    ["Titulo ENORME en la portada - Vive conectado al nuevo San Miguel", "Titulo Principal", "MUY GRANDE (54 px)", "MUY GRUESA (Bold)", "En la foto grande de la portada"],
    ["Subtitulo de cada seccion - Invierte en San Miguel", "Titulo de Seccion", "Bastante grande (48-56 px)", "MUY GRUESA (Bold)", "Encima de cada seccion importante"],
    ["Titulos medianos - Rentabilidad bruta estimada", "Titulo Pequeno", "Mediano (24 px)", "Gruesa (SemiBold)", "En las tarjetas de informacion"],
    ["Texto explicativo normal - parrafos largos", "Parrafo de Descripcion", "Normal (16 px)", "Delgada (Regular)", "En los textos largos que explican cosas"],
    ["Texto de contenido importante", "Parrafo Principal", "Normal-Mediano (16-18 px)", "Un poco gruesa (Medium)", "En parrafos importantes de la pagina"],
    ["Texto de Nombre, Apellido en formularios", "Etiqueta de Formulario", "Pequeno (14-15 px)", "Un poco gruesa (Medium)", "Encima de los campos donde escribes"],
    ["Letras muy pequenas - aclaraciones, notas", "Texto Pequeno", "Muy pequeno (12-14 px)", "Delgada (Regular)", "En descripciones cortas y notas"],
    ["Cotizar ahora dentro de los botones verdes", "Texto en Botones", "Normal (16 px)", "Gruesa (SemiBold)", "Dentro de los botones de accion"],
    ["Los numeros grandes - 95 Departamentos, 62 Estacionamientos", "Numeros Grandes/Metricas", "MUY GRANDE (32-40 px)", "MUY GRUESA (Bold)", "En las tarjetas de estadisticas"],
    ["UF 3.250 o $97.500.000 en los departamentos", "Precio", "Mediano (17-22 px)", "MUY GRUESA (Bold)", "Al lado de cada departamento"],
    ["Los links del menu de arriba - Inicio, Ubicacion, Proyecto", "Menu de Navegacion", "Pequeno (14 px)", "Un poco gruesa (Medium)", "En la barra de navegacion de arriba"],
    ["Palabras subrayadas o que parecen clicables", "Enlaces/Links", "Pequeno-Normal (14-16 px)", "Un poco gruesa (Medium)", "Distribuidos en la pagina"],
    ["Texto al pie de la pagina - informacion legal", "Pie de Pagina (Footer)", "Pequeno (14 px)", "Delgada (Regular)", "Al final de la pagina"],
    ["Depto. 301 - Piso 3, Depto. 302 - Piso 3", "Nombre del Departamento", "Mediano (17-20 px)", "MUY GRUESA (Bold)", "En cada tarjeta de departamento"],
    ["3D/2B, 120m2, 89m2 util", "Caracteristicas del Departamento", "Pequeno (14 px)", "Delgada (Regular)", "Bajo el nombre del departamento"],
    ["Titulos dentro de ventanas emergentes", "Titulos en Ventanas Emergentes", "Mediano (24 px)", "MUY GRUESA (Bold)", "Cuando sale una ventana con detalles"],
    ["Texto normal dentro de ventanas emergentes", "Texto en Ventanas Emergentes", "Normal (16 px)", "Delgada (Regular)", "En las ventanas emergentes"],
]

for row_idx, row_data in enumerate(data, 8):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.font = Font(name="Nunito Sans", size=9)
        cell.border = border
        cell.alignment = wrap_align
        ws.row_dimensions[row_idx].height = 45

# RESUMEN FINAL
ws['A26'] = "RESUMEN SIMPLE"
ws['A26'].font = title_font
ws['A26'].fill = title_fill
ws['A26'].alignment = center_align
ws.merge_cells('A26:E26')
ws.row_dimensions[26].height = 20

resumen = [
    "SOLO SE USA UNA LETRA: Nunito Sans (nada de Arial, Helvetica, Times, etc.)",
    "SOLO 4 GROSORES: Delgada (400), Un poco gruesa (500), Gruesa (600), Muy gruesa (700)",
    "TAMANOS VARIAN: Desde 12 pixeles (muy pequeno) hasta 54 pixeles (gigante)",
    "TODO SE ADAPTA: En telefono se ven mas pequenos, en computadora mas grandes",
    "CONSISTENCIA: Toda la pagina usa la misma letra, es profesional y limpio",
]

for idx, punto in enumerate(resumen, 27):
    ws[f'A{idx}'] = punto
    ws[f'A{idx}'].font = Font(name="Nunito Sans", size=10, bold=True)
    ws.merge_cells(f'A{idx}:E{idx}')
    ws.row_dimensions[idx].height = 18

wb.save('Tipografia_Boulevard_Santa_Rosa.xlsx')
print("Excel creado exitosamente!")
